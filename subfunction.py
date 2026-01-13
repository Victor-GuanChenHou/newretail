def outputexcel():
    import pyodbc
    import pandas as pd
    from openpyxl import Workbook
    import pandas as pd
    from datetime import datetime
    from openpyxl.styles import PatternFill
    import os
    from openpyxl import Workbook
    from dotenv import load_dotenv
    import os
    ENV = './.env' 
    load_dotenv(dotenv_path=ENV)
    DB_host = os.getenv('DB_host')
    DB_password = os.getenv('DB_password')
    DB_uid=os.getenv('DB_uid')
    DATABASE=os.getenv('DATABASE')
    # 讀取進價.xlsx文件
    file_path = '進價.xlsx'  # 請根據實際文件位置修改

    # 使用pandas的read_excel函數讀取Excel文件
    inputdata = pd.read_excel(file_path, engine='openpyxl')
    try:
        # 建立與 SQL Server 的連線
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={DB_host};"  # 替換為完整的伺服器名稱
            f"DATABASE={DATABASE};"
            f"UID={DB_uid};"  # 使用 sa 作為使用者名稱
            f"PWD={DB_password};"  # 替換為 sa 的密碼
            "Trusted_Connection=no;" # 明確使用 SQL Server 認證
              
        )

        # 用戶選擇的品牌（假設這來自某個介面） 
        CombMO002_ItemIndex = 'S02012'

        # SQL1 查詢語句
        Sql1 = """
            SELECT PLU_ID, PLU_NAME, PLU_SPEC, BIN_NAME, BIN_DESC, LABEL_F, UNIT, EXP_DATE, QTY, BOOKING, AVAILABLE
            FROM View_WMS02
            WHERE 1=1
        """

        # 根據品牌選擇拼接 Sql2
        if CombMO002_ItemIndex == "0":
            Sql2 = ""
        else:
            MA076 = CombMO002_ItemIndex
            Sql2 = """
                AND RTRIM(PLU_ID) COLLATE Chinese_Taiwan_Stroke_BIN IN (
                    SELECT RTRIM(MI002)
                    FROM COPMA
                    INNER JOIN LKFMJ ON MJ001 = MA001
                    LEFT JOIN LKFMH ON MH001 = MJ002
                    LEFT JOIN LKFMI ON MI001 = MH001
                    WHERE 1=1
                    AND MA076 = ?
                    UNION
                    SELECT RTRIM(MD002) AS MI002
                    FROM COPMA
                    INNER JOIN LKFMG ON MG001 = MA001
                    LEFT JOIN LKFMC ON MC001 = MG002
                    LEFT JOIN LKFMD ON MD001 = MC001
                    WHERE 1=1
                    AND MA076 = ?
                )
            """


        # 完整的查詢語句
        Sql3 =   " ORDER BY PLU_ID, BIN_NAME, EXP_DATE, LABEL_F"

        # 最終的 SQL 查詢語句
        Sql = Sql1 + Sql2 + Sql3

        # 使用 pandas 執行查詢並傳遞參數以防止 SQL 注入
        df = pd.read_sql(Sql, conn, params=(MA076, MA076))
    except:
        message='連線失敗'
        
        return message
        
    df.columns =['品號','品名','規格','標的代號','標的名稱','貨品標籤','單位','有效日期','數量','鎖定數量','可用數量']

    data={}
    shelflifeday=[]
    day_1=[]
    day_2=[]
    day_3=[]
    day_4=[]
    shelflifeday_2=[]
    end=[]
    data['品號']=df['品號']
    data['品名']=df['品名']

    for z in range(len(data['品號'])):
        filtered_df = inputdata[inputdata['品號'] == data['品號'][z]]
        print("欄位名稱：", filtered_df.columns.tolist())
        if filtered_df.empty:
            shelflifeday.append(0)
        else:
            shelflifeday.append(filtered_df['保存期限(天)'].iloc[0])
    data['保存期限天數']=shelflifeday
    data['允收期限2/3'] = [2/3] * len(df['品名'])
    data['允收期限1/2'] = [1/2] * len(df['品名'])
    for z in range(len(data['品號'])):
        if data['保存期限天數'][z] != 0:
            day_1.append(int(data['保存期限天數'][z])*2/3)
            day_2.append(int(data['保存期限天數'][z])*1/2)
        else:
            day_1.append(0)
            day_2.append(0)
        str_1=df['有效日期'][z]
        str_2=f"{str_1[:4]}/{str_1[4:6]}/{str_1[6:]}"
        target_date = datetime.strptime(str_2, '%Y/%m/%d')
        shelflifeday_2.append(str_2)
        end.append( (target_date - datetime.today()).days)
    
    data['計算天數2/3']=day_1
    data['計算天數1/2']=day_2
    data['可用數量']=df['可用數量']
    data['有效日期']=shelflifeday_2
    data['結果']=end
    for z in range(len(data['品號'])):
        if data['計算天數2/3'][z]==0:
            day_3.append(None)
            day_4.append(None)
        else:
            day_3.append(int(data['結果'][z]-data['計算天數2/3'][z]))
            day_4.append(int(data['結果'][z]-data['計算天數1/2'][z]))
    data['1/3允收餘天數']=day_3
    data['1/2允收餘天數']=day_4
    data=pd.DataFrame(data)
    data.columns=['品號','品名','保存期限天數','允收期限2/3','允收期限1/2','計算天數2/3','計算天數1/2','可用數量','有效日期','結果','1/3允收餘天數','1/2允收餘天數']

    merged_data = data.groupby(["品號", "有效日期"], as_index=False).agg({
        "品名": "first",
        "保存期限天數": "first",
        "允收期限2/3": "first",
        "允收期限1/2": "first",
        "計算天數2/3": "first",
        "計算天數1/2": "first",
        "可用數量": "sum" ,
        "有效日期": "first",
        "結果": "first",
        "1/3允收餘天數": "first",
        "1/2允收餘天數": "first"
        
    })
    
    data=merged_data
    
    
    # 根據 "品號" 進行分組，並對 "可用數量" 求和
    df_grouped = df.groupby(['品號','品名'],    as_index=False)['可用數量'].sum()
    df_grouped.columns=['品號','品名','總和']
    # 重新設置索引，將總和欄位插入最後

    # 創建 Excel 工作簿
    wb = Workbook()

    # 添加工作表
    ws = wb.active
    ws.title = "王座青田明細"
    header_fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")

    # 在 A 到 K 欄顯示 df 的欄位名稱和數據
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    # 將 df 輸出到 A 到 K 欄
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    ##################################
    for col_idx, col_name in enumerate(df_grouped.columns, start=13):  # M 欄開始
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    # 將 df_grouped 輸出到 M 到 O 欄
    for r_idx, row in df_grouped.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 13, value=value)
    #############################
    for col_idx, col_name in enumerate(data.columns, start=18):  
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    for r_idx, row in data.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 18, value=value)

    # 儲存 Excel 文件
    

    # 要儲存的檔案名稱
    filename = "王座青田明細.xlsx"
    full_path = os.path.join(filename)

    
    wb.save(full_path)

    # output_file = "test.xlsx"
    # df.to_excel(output_file, index=False, engine='openpyxl')

    # 關閉連線
    conn.close()
    message=full_path
    return message
def fullday(path):
    import pyodbc
    import pandas as pd
    from openpyxl import Workbook
    import pandas as pd
    from datetime import datetime
    from openpyxl.styles import PatternFill
    import os
    from openpyxl import Workbook
    
    # 讀取進價.xlsx文件
    file_path = '進價.xlsx'  # 請根據實際文件位置修改

    # 使用pandas的read_excel函數讀取Excel文件
    inputdata = pd.read_excel(file_path, engine='openpyxl')
    df=pd.read_excel(path, engine='openpyxl')
    data={}
    shelflifeday=[]
    day_1=[]
    day_2=[]
    day_3=[]
    day_4=[]
    shelflifeday_2=[]
    end=[]
    data['品號']=df['商品編號']
    data['品名']=df['商品名稱']

    for z in range(len(data['品號'])):
        filtered_df = inputdata[inputdata['品號'] == data['品號'][z]]
        if filtered_df.empty:
            shelflifeday.append(0)
        else:
            shelflifeday.append(filtered_df['保存期限(天)'].iloc[0])
    data['保存期限天數']=shelflifeday
    data['允收期限2/3'] = [2/3] * len(df['商品名稱'])
    data['允收期限1/2'] = [1/2] * len(df['商品名稱'])
    for z in range(len(data['品號'])):
        if data['保存期限天數'][z] != 0:
            day_1.append(int(data['保存期限天數'][z])*2/3)
            day_2.append(int(data['保存期限天數'][z])*1/2)
        else:
            day_1.append(0)
            day_2.append(0)
        str_1=df['有效日期'][z]
        
        shelflifeday_2.append(str_1)
        target_date = datetime.strptime(str_1, '%Y/%m/%d')
        end.append( (target_date - datetime.today()).days)
    
    data['計算天數2/3']=day_1
    data['計算天數1/2']=day_2
    data['可用數量']=df['庫存數量']
    data['有效日期']=shelflifeday_2
    data['結果']=end
    for z in range(len(data['品號'])):
        if data['計算天數2/3'][z]==0:
            day_3.append(None)
            day_4.append(None)
        else:
            day_3.append(int(data['結果'][z]-data['計算天數2/3'][z]))
            day_4.append(int(data['結果'][z]-data['計算天數1/2'][z]))
    data['1/3允收餘天數']=day_3
    data['1/2允收餘天數']=day_4
    data=pd.DataFrame(data)
    data.columns=['品號','品名','保存期限天數','允收期限2/3','允收期限1/2','計算天數2/3','計算天數1/2','可用數量','有效日期','結果','1/3允收餘天數','1/2允收餘天數']
    merged_data = data.groupby(["品號", "有效日期"], as_index=False).agg({
        "品名": "first",
        "保存期限天數": "first",
        "允收期限2/3": "first",
        "允收期限1/2": "first",
        "計算天數2/3": "first",
        "計算天數1/2": "first",
        "可用數量": "sum" ,
        "有效日期": "first",
        "結果": "first",
        "1/3允收餘天數": "first",
        "1/2允收餘天數": "first"
        
    })
    data=merged_data
    
    # 根據 "品號" 進行分組，並對 "可用數量" 求和
    
    # 重新設置索引，將總和欄位插入最後

    # 創建 Excel 工作簿
    wb = Workbook()

    # 添加工作表
    ws = wb.active
    ws.title = "王座全日明細"
    header_fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")

    # 在 A 到 K 欄顯示 df 的欄位名稱和數據
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    # 將 df 輸出到 A 到 K 欄
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    
    #############################
    for col_idx, col_name in enumerate(data.columns, start=21):  
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    for r_idx, row in data.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 21, value=value)

    # 儲存 Excel 文件
    

    # 要儲存的檔案名稱
    filename = "王座全日明細.xlsx"
    full_path = os.path.join(filename)

    
    wb.save(full_path)

    
    
    message=full_path
    return message
def ALL(path):
    import pyodbc
    import pandas as pd
    from openpyxl import Workbook
    import pandas as pd
    from datetime import datetime, timedelta
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    import os
    from openpyxl import Workbook
    from dotenv import load_dotenv
    import os
    ENV = './.env' 
    load_dotenv(dotenv_path=ENV)
    DB_host = os.getenv('DB_host')
    DB_password = os.getenv('DB_password')
    DB_uid=os.getenv('DB_uid')
    DATABASE=os.getenv('DATABASE')
    # 讀取進價.xlsx文件
    file_path = '進價.xlsx'  # 請根據實際文件位置修改

    # 使用pandas的read_excel函數讀取Excel文件
    inputdata = pd.read_excel(file_path, engine='openpyxl')
    try:
        # 建立與 SQL Server 的連線
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={DB_host};"  # 替換為完整的伺服器名稱
            f"DATABASE={DATABASE};"
            f"UID={DB_uid};"  # 使用 sa 作為使用者名稱
            f"PWD={DB_password};"  # 替換為 sa 的密碼
            "Trusted_Connection=no;" # 明確使用 SQL Server 認證
              
        )

        # 用戶選擇的品牌（假設這來自某個介面） 
        CombMO002_ItemIndex = 'S02012'

        # SQL1 查詢語句
        Sql1 = """
            SELECT PLU_ID, PLU_NAME, PLU_SPEC, BIN_NAME, BIN_DESC, LABEL_F, UNIT, EXP_DATE, QTY, BOOKING, AVAILABLE
            FROM View_WMS02
            WHERE 1=1
        """

        # 根據品牌選擇拼接 Sql2
        if CombMO002_ItemIndex == "0":
            Sql2 = ""
        else:
            MA076 = CombMO002_ItemIndex
            Sql2 = """
                AND RTRIM(PLU_ID) COLLATE Chinese_Taiwan_Stroke_BIN IN (
                    SELECT RTRIM(MI002)
                    FROM COPMA
                    INNER JOIN LKFMJ ON MJ001 = MA001
                    LEFT JOIN LKFMH ON MH001 = MJ002
                    LEFT JOIN LKFMI ON MI001 = MH001
                    WHERE 1=1
                    AND MA076 = ?
                    UNION
                    SELECT RTRIM(MD002) AS MI002
                    FROM COPMA
                    INNER JOIN LKFMG ON MG001 = MA001
                    LEFT JOIN LKFMC ON MC001 = MG002
                    LEFT JOIN LKFMD ON MD001 = MC001
                    WHERE 1=1
                    AND MA076 = ?
                )
            """


        # 完整的查詢語句
        Sql3 =   " ORDER BY PLU_ID, BIN_NAME, EXP_DATE, LABEL_F"

        # 最終的 SQL 查詢語句
        Sql = Sql1 + Sql2 + Sql3

        # 使用 pandas 執行查詢並傳遞參數以防止 SQL 注入
        df = pd.read_sql(Sql, conn, params=(MA076, MA076))
    except:
        message='連線失敗'
        
        return message
        
    df.columns =['品號','品名','規格','標的代號','標的名稱','貨品標籤','單位','有效日期','數量','鎖定數量','可用數量']

    data={}
    shelflifeday=[]
    day_1=[]
    day_2=[]
    day_3=[]
    day_4=[]
    shelflifeday_2=[]
    end=[]
    data['品號']=df['品號']
    data['品名']=df['品名']

    for z in range(len(data['品號'])):
        filtered_df = inputdata[inputdata['品號'] == data['品號'][z]]
        if filtered_df.empty:
            shelflifeday.append(0)
        else:
            shelflifeday.append(filtered_df['保存期限(天)'].iloc[0])
    data['保存期限天數']=shelflifeday
    data['允收期限2/3'] = [2/3] * len(df['品名'])
    data['允收期限1/2'] = [1/2] * len(df['品名'])
    for z in range(len(data['品號'])):
        if data['保存期限天數'][z] != 0:
            day_1.append(int(data['保存期限天數'][z])*2/3)
            day_2.append(int(data['保存期限天數'][z])*1/2)
        else:
            day_1.append(0)
            day_2.append(0)
        str_1=df['有效日期'][z]
        str_2=f"{str_1[:4]}/{str_1[4:6]}/{str_1[6:]}"
        target_date = datetime.strptime(str_2, '%Y/%m/%d')
        shelflifeday_2.append(str_2)
        end.append( (target_date - datetime.today()).days)
    
    data['計算天數2/3']=day_1
    data['計算天數1/2']=day_2
    data['可用數量']=df['可用數量']
    data['有效日期']=shelflifeday_2
    data['結果']=end
    for z in range(len(data['品號'])):
        if data['計算天數2/3'][z]==0:
            day_3.append(None)
            day_4.append(None)
        else:
            day_3.append(int(data['結果'][z]-data['計算天數2/3'][z]))
            day_4.append(int(data['結果'][z]-data['計算天數1/2'][z]))
    data['1/3允收餘天數']=day_3
    data['1/2允收餘天數']=day_4
    data=pd.DataFrame(data)
    data.columns=['品號','品名','保存期限天數','允收期限2/3','允收期限1/2','計算天數2/3','計算天數1/2','可用數量','有效日期','結果','1/3允收餘天數','1/2允收餘天數']

    merged_data = data.groupby(["品號", "有效日期"], as_index=False).agg({
        "品名": "first",
        "保存期限天數": "first",
        "允收期限2/3": "first",
        "允收期限1/2": "first",
        "計算天數2/3": "first",
        "計算天數1/2": "first",
        "可用數量": "sum" ,
        "有效日期": "first",
        "結果": "first",
        "1/3允收餘天數": "first",
        "1/2允收餘天數": "first"
        
    })
    cingting_grouped = df.groupby(['品號','品名'],    as_index=False)['可用數量'].sum()
    cingting_grouped.columns=['品號','品名','總和']
    data=merged_data
    cingting_o=df
    
    cingting_day=data


    ##########全日##########
    df=pd.read_excel(path, engine='openpyxl')
    data={}
    shelflifeday=[]
    day_1=[]
    day_2=[]
    day_3=[]
    day_4=[]
    shelflifeday_2=[]
    end=[]
    data['品號']=df['商品編號']
    data['品名']=df['商品名稱']

    for z in range(len(data['品號'])):
        filtered_df = inputdata[inputdata['品號'] == data['品號'][z]]
        if filtered_df.empty:
            shelflifeday.append(0)
        else:
            shelflifeday.append(filtered_df['保存期限(天)'].iloc[0])
    data['保存期限天數']=shelflifeday
    data['允收期限2/3'] = [2/3] * len(df['商品名稱'])
    data['允收期限1/2'] = [1/2] * len(df['商品名稱'])
    for z in range(len(data['品號'])):
        if data['保存期限天數'][z] != 0:
            day_1.append(int(data['保存期限天數'][z])*2/3)
            day_2.append(int(data['保存期限天數'][z])*1/2)
        else:
            day_1.append(0)
            day_2.append(0)
        str_1=df['有效日期'][z]
        
        shelflifeday_2.append(str_1)
        target_date = datetime.strptime(str_1, '%Y/%m/%d')
        end.append( (target_date - datetime.today()).days)
    
    data['計算天數2/3']=day_1
    data['計算天數1/2']=day_2
    data['可用數量']=df['庫存數量']
    data['有效日期']=shelflifeday_2
    data['結果']=end
    for z in range(len(data['品號'])):
        if data['計算天數2/3'][z]==0:
            day_3.append(None)
            day_4.append(None)
        else:
            day_3.append(int(data['結果'][z]-data['計算天數2/3'][z]))
            day_4.append(int(data['結果'][z]-data['計算天數1/2'][z]))
    data['1/3允收餘天數']=day_3
    data['1/2允收餘天數']=day_4
    data=pd.DataFrame(data)
    data.columns=['品號','品名','保存期限天數','允收期限2/3','允收期限1/2','計算天數2/3','計算天數1/2','可用數量','有效日期','結果','1/3允收餘天數','1/2允收餘天數']
    merged_data = data.groupby(["品號", "有效日期"], as_index=False).agg({
        "品名": "first",
        "保存期限天數": "first",
        "允收期限2/3": "first",
        "允收期限1/2": "first",
        "計算天數2/3": "first",
        "計算天數1/2": "first",
        "可用數量": "sum" ,
        "有效日期": "first",
        "結果": "first",
        "1/3允收餘天數": "first",
        "1/2允收餘天數": "first"
        
    })
    data=merged_data
    changl_o=df
    changl_day=data

    # 在 A 到 K 欄顯示 df 的欄位名稱和數據
    # 創建 Excel 工作簿
    wb = Workbook()

    # 添加工作表
    ws = wb.active
    ws.title = "總表"
    header_fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")
    
    #############總表################
   
    ##############青田################
    start_month = datetime.today().replace(day=1)

    months = [(start_month + pd.DateOffset(months=i)).strftime('%Y%m') for i in range(18)]
    months.append('超過18個月')
    months.append('已過期')
    months.insert(0, '品號')
    months.insert(1, '青田品名')
    months.insert(2, '規格')
    months.insert(3, '成本')
    months.append('總計')
    months.append('總成本(未稅)')
    # 初始化結果表格
    result = pd.DataFrame(columns=months)
    
    # 填充缺失值
    result = result.fillna(0)
    
    # 根據 cingting_day 中的「品號」去掉重複並填充到 result 的「品號」欄位
    result['品號'] = cingting_day['品號'].drop_duplicates().reset_index(drop=True)
   
    
    date=datetime.strptime(cingting_day['有效日期'][0], '%Y/%m/%d')
    date = date.strftime('%Y%m')   
    mm=(start_month + pd.DateOffset(months=17)).strftime('%Y%m')
    start_month_str=start_month.strftime('%Y%m')
    
    for z in range(len(result['品號'])):
        for p in range(len(cingting_day['品號'])):
            if result['品號'][z] == cingting_day['品號'][p]:
                
                date=datetime.strptime(cingting_day['有效日期'][p], '%Y/%m/%d')
                date_str = date.strftime('%Y%m') 
                if int(mm)>=int(date_str) and int(start_month_str)<=int(date_str):
                    if pd.isna(result[date_str][z]):
                        result.loc[z,date_str ] = cingting_day.loc[p, '可用數量']
                    else:
                        result.loc[z,date_str] = cingting_day.loc[p, '可用數量']+result[date_str][z]
                elif int(mm)<int(date_str):
                    if pd.isna(result['超過18個月'][z]):
                        result.loc[z,'超過18個月' ] = cingting_day.loc[p, '可用數量']
                    else:
                        result.loc[z,'超過18個月' ] = cingting_day.loc[p, '可用數量']+result['超過18個月'][z]
                else :
                    if pd.isna(result['已過期'][z]):
                        result.loc[z,'已過期' ] = cingting_day.loc[p, '可用數量']
                    else:
                        result.loc[z,'已過期' ] = cingting_day.loc[p, '可用數量']+result['已過期'][z]

    for z in range(len(result['品號'])):
        for p in range(len(cingting_o['品號'])):
            if result['品號'][z] == cingting_o['品號'][p]:
                result.loc[z, '青田品名'] = cingting_o.loc[p, '品名']
                result.loc[z, '規格'] = cingting_o.loc[p, '規格']
                
        if pd.isna(result['總計'][z]):
            result.loc[z,'總計'] = 0
        for k in range(18):
            if pd.isna(result[(start_month + pd.DateOffset(months=k)).strftime('%Y%m') ][z]):
                data_result=0
            else:
                data_result=result[(start_month + pd.DateOffset(months=k)).strftime('%Y%m')][z]
            result.loc[z,'總計']=data_result+result['總計'][z]
        if pd.isna(result['超過18個月'][z]):
            data_result=0
        else:
            data_result=result['超過18個月'][z]
        result.loc[z,'總計']=data_result+result['總計'][z]     
        if pd.isna(result['已過期'][z]):
            data_result=0
        else:
            data_result=result['已過期'][z]
        result.loc[z,'總計']=data_result+result['總計'][z]
        for t in range(len(inputdata)):
            if result['品號'][z]==inputdata['品號'][t]:
                result.loc[z,'成本']=inputdata['成本'][t]   
                result.loc[z,'總成本(未稅)']=inputdata['成本'][t]*result['總計'][z]
    start_month = datetime.today().replace(day=1)
    months = [(start_month + pd.DateOffset(months=i)).strftime('%Y%m') for i in range(18)]
    new_row = {
            '品號': None, 
            '青田品名': None, 
            '規格': None, 
            '成本': '總計', 
            **{month: result[month].sum() for month in months},  # Generate columns for the 18 months
            '超過18個月': result['超過18個月'].sum(), 
            '總計': result['總計'].sum(), 
            '已過期': result['已過期'].sum(), 
            '總成本(未稅)': result['總成本(未稅)'].sum()
        }
    new_row_df = pd.DataFrame([new_row])
    result = pd.concat([result, new_row_df], ignore_index=True)    
    #匯總數據
    cingting_result=result


    #################全日####################
    start_month = datetime.today().replace(day=1)

    months = [(start_month + pd.DateOffset(months=i)).strftime('%Y%m') for i in range(18)]
    
    months.append('超過18個月')
    months.append('已過期')
    months.insert(0, '品號')
    months.insert(1, '全日品名')
    months.insert(2, '規格')
    months.insert(3, '成本')
    months.append('總計')
    months.append('總成本(未稅)')

    # 初始化結果表格
    result = pd.DataFrame(columns=months)
    
    # 填充缺失值
    result = result.fillna(0)
    
    # 根據 changl_day 中的「品號」去掉重複並填充到 result 的「品號」欄位
    result['品號'] = changl_day['品號'].drop_duplicates().reset_index(drop=True)
   
    
    date=datetime.strptime(changl_day['有效日期'][0], '%Y/%m/%d')
    date = date.strftime('%Y%m')   
    mm=(start_month + pd.DateOffset(months=17)).strftime('%Y%m')
    start_month_str=start_month.strftime('%Y%m')
    
    for z in range(len(result['品號'])):
        for p in range(len(changl_day['品號'])):
            if result['品號'][z] == changl_day['品號'][p]:
                
                date=datetime.strptime(changl_day['有效日期'][p], '%Y/%m/%d')
                date_str = date.strftime('%Y%m') 
                
                if int(mm)>=int(date_str) and int(start_month_str)<=int(date_str):
                    if pd.isna(result[date_str][z]):
                        m_data=0
                    else:
                        m_data=result[date_str][z]
                    result.loc[z,date_str ] = changl_day.loc[p, '可用數量']+m_data
                elif int(mm)<int(date_str):
                    if pd.isna(result['超過18個月'][z]):
                        m_data=0
                    else:
                        m_data=result['超過18個月'][z]
                    result.loc[z,'超過18個月' ] = changl_day.loc[p, '可用數量']+m_data
                else :
                    if pd.isna(result['已過期'][z]):
                        m_data=0
                    else:
                        m_data=result['已過期'][z]
                    result.loc[z,'已過期' ] = changl_day.loc[p, '可用數量']+m_data
    for z in range(len(result['品號'])):
        for p in range(len(changl_o['商品編號'])):
            if result['品號'][z] == changl_o['商品編號'][p]:
                result.loc[z, '全日品名'] = changl_o.loc[p, '商品名稱']
                
        if pd.isna(result['總計'][z]):
            result.loc[z,'總計'] = 0
        for k in range(18):
            if pd.isna(result[(start_month + pd.DateOffset(months=k)).strftime('%Y%m') ][z]):
                data_result=0
            else:
                data_result=result[(start_month + pd.DateOffset(months=k)).strftime('%Y%m')][z]
            result.loc[z,'總計']=data_result+result['總計'][z]
        if pd.isna(result['超過18個月'][z]):
            data_result=0
        else:
            data_result=result['超過18個月'][z]
        result.loc[z,'總計']=data_result+result['總計'][z]     
        if pd.isna(result['已過期'][z]):
            data_result=0
        else:
            data_result=result['已過期'][z]
        result.loc[z,'總計']=data_result+result['總計'][z]
        for t in range(len(inputdata)):
            if result['品號'][z]==inputdata['品號'][t]:
                
                result.loc[z,'成本']=inputdata['成本'][t]   
                result.loc[z,'總成本(未稅)']=inputdata['成本'][t]*result['總計'][z]
        # for j in range(len(cingting_o)):
        #     if result['品號'][z]==cingting_o['品號'][j]:
        #         result.loc[z, '規格'] = cingting_o[ '規格'][j]
        #         print(cingting_o['品號'][j])
        success,result.loc[z, '規格']=get_format(result['品號'][z])

    start_month = datetime.today().replace(day=1)
    months = [(start_month + pd.DateOffset(months=i)).strftime('%Y%m') for i in range(18)]
    new_row = {
            '品號': None, 
            '全日品名': None, 
            '規格': None, 
            '成本': '總計', 
            **{month: result[month].sum() for month in months},  # Generate columns for the 18 months
            '超過18個月': result['超過18個月'].sum(), 
            '總計': result['總計'].sum(), 
            '已過期': result['已過期'].sum(), 
            '總成本(未稅)': result['總成本(未稅)'].sum()
        }
    new_row_df = pd.DataFrame([new_row])
    result = pd.concat([result, new_row_df], ignore_index=True)    
    #匯總數據
    changl_result=result
    col=['品號','品名','總計']
    pinghao_a = pd.DataFrame(columns=col)
    pinghao_b = pd.DataFrame(columns=col)
    for i in range(len(cingting_result['品號'])):
        
        if pd.isna(cingting_result['品號'][i]):
            pass
        else:
            if cingting_result['品號'][i].startswith(("CP", "KO")):
                pinghao_b.loc[i] = [
                    cingting_result['品號'][i],
                    cingting_result['青田品名'][i],
                    cingting_result['總計'][i]
                ]
            else:
                pinghao_a.loc[i] = [
                    cingting_result['品號'][i],
                    cingting_result['青田品名'][i],
                    cingting_result['總計'][i]
                ]
                
    pinghao_a.reset_index(drop=True, inplace=True)
    pinghao_b.reset_index(drop=True, inplace=True)
    pinghao_c = pd.DataFrame(columns=col)
    pinghao_d = pd.DataFrame(columns=col)
    for i in range(len(changl_result['品號'])):
        
        if pd.isna(changl_result['品號'][i]):
            pass
        else:
            if changl_result['品號'][i].startswith(("CP", "KO")):
                pinghao_d.loc[i] = [
                    changl_result['品號'][i],
                    changl_result['全日品名'][i],
                    changl_result['總計'][i]
                ]
                
            else:
                
                pinghao_c.loc[i] = [
                    changl_result['品號'][i],
                    changl_result['全日品名'][i],
                    changl_result['總計'][i]
                ]
    pinghao_c.reset_index(drop=True, inplace=True)
    pinghao_d.reset_index(drop=True, inplace=True)
    
    for index, row in pinghao_c.iterrows():
        品號_value = row['品號']
        品名_value = row['品名']
        總計_value = row['總計']
            
            # 判斷是否已存在
        if 品號_value in pinghao_a['品號'].values:
                # 更新總計
            pinghao_a.loc[pinghao_a['品號'] == 品號_value, '總計'] += 總計_value
        else:
            # 不存在則新增
            pinghao_a = pd.concat([pinghao_a, pd.DataFrame({"品號": [品號_value],"品名": [品名_value], "總計": [總計_value]})], ignore_index=True)
    for index, row in pinghao_d.iterrows():
        品號_value = row['品號']
        品名_value = row['品名']
        總計_value = row['總計']
            
            # 判斷是否已存在
        if 品號_value in pinghao_b['品號'].values:
                # 更新總計
            
            pinghao_b.loc[pinghao_b['品號'] == 品號_value, '總計'] += 總計_value
        else:
                # 不存在則新增
            
            pinghao_b = pd.concat([pinghao_b, pd.DataFrame({"品號": [品號_value],"品名": [品名_value], "總計": [總計_value]})], ignore_index=True)

    new_row = {
            '品號': None, 
            '品名': '總計', 
            '總計':pinghao_a['總計'].sum()
            
        }
    pinghao_a_df = pd.DataFrame([new_row])
    pinghao_a = pd.concat([pinghao_a, pinghao_a_df], ignore_index=True)                   
    
    new_row = {
            '品號': None, 
            '品名': '總計', 
            '總計':pinghao_b['總計'].sum()
            
        }
    pinghao_b_df = pd.DataFrame([new_row])
    pinghao_b = pd.concat([pinghao_b, pinghao_b_df], ignore_index=True)   
    length=len(cingting_day['品號'])+3  
    length_2=len(cingting_result['品號'])+3  
    length_3=len(pinghao_a)+3  
    
    ###############最上總和資料######################
    todaystr=datetime.today().strftime('%Y%m%d')
    alldata={
        todaystr+'倉庫週庫存':todaystr+'倉庫週庫存',
        '1':todaystr+'倉庫週庫存1',
        '2':todaystr+'倉庫週庫存2',
        '月份統計表':'電商總計數量/金額',
        **{month: changl_result[month][len(changl_result[month])-1]+ cingting_result[month][len(cingting_result[month])-1]for month in months},
        '超過18個月': changl_result['超過18個月'][len(changl_result['超過18個月'])-1]+ cingting_result['超過18個月'][len(cingting_result['超過18個月'])-1], 
        '已過期': changl_result['已過期'][len(changl_result['已過期'])-1]+ cingting_result['已過期'][len(cingting_result['已過期'])-1], 
        '總計': changl_result['總計'][len(changl_result['總計'])-1]+ cingting_result['總計'][len(cingting_result['總計'])-1], 
        '總成本(未稅)': changl_result['總成本(未稅)'][len(changl_result['總成本(未稅)'])-1]+ cingting_result['總成本(未稅)'][len(cingting_result['總成本(未稅)'])-1]
    }
    alldata = pd.DataFrame([alldata])
    for col_idx, col_name in enumerate(alldata.columns, start=1):
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        
    for r_idx, row in alldata.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    ws.merge_cells('A1:C2')
    for col_idx, col_name in enumerate(cingting_result.columns, start=1):
        cell=ws.cell(row=6, column=col_idx, value=col_name)  # 放置欄位名稱
        column_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.fill = column_fill  
    for r_idx, row in cingting_result.iterrows():
        for c_idx, value in enumerate(row):
            cell=ws.cell(row=r_idx + 7, column=c_idx + 1, value=value)
            if(r_idx%2==0):
                column_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            else:
                column_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill=column_fill
    for col_idx, col_name in enumerate(changl_result.columns, start=1):
        cell=ws.cell(row=6+length_2, column=col_idx, value=col_name)  # 放置欄位名稱
        column_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        cell.fill = column_fill  
    for r_idx, row in changl_result.iterrows():
        for c_idx, value in enumerate(row):
            cell=ws.cell(row=r_idx + 7+length_2, column=c_idx + 1, value=value)
            if(r_idx%2==0):
                column_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            else:
                column_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill=column_fill
    for col_idx, col_name in enumerate(cingting_day.columns, start=29):  
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    for r_idx, row in cingting_day.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 29, value=value)
    for col_idx, col_name in enumerate(changl_day.columns, start=29):  
        cell=ws.cell(row=length, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    for r_idx, row in changl_day.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + length+1, column=c_idx + 29, value=value)
    
    for col_idx, col_name in enumerate(pinghao_a.columns, start=45):  
        cell=ws.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        column_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.fill = column_fill  
    for r_idx, row in pinghao_a.iterrows():
        for c_idx, value in enumerate(row):
            cell=ws.cell(row=r_idx +2, column=c_idx + 45, value=value)
            if(r_idx%2==0):
                column_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            else:
                column_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill=column_fill
    for col_idx, col_name in enumerate(pinghao_b.columns, start=45):  
        cell=ws.cell(row=length_3, column=col_idx, value=col_name)  # 放置欄位名稱
        column_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.fill = column_fill  
    for r_idx, row in pinghao_b.iterrows():
        for c_idx, value in enumerate(row):
            cell=ws.cell(row=r_idx+length_3 +1, column=c_idx + 45, value=value)
            if(r_idx%2==0):
                column_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            else:
                column_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill=column_fill

    

    #######################青田###############################
    ws_cingting = wb.create_sheet("王座青田明細")
    for col_idx, col_name in enumerate(cingting_o.columns, start=1):
        cell=ws_cingting.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    # 將 df 輸出到 A 到 K 欄
    for r_idx, row in cingting_o.iterrows():
        for c_idx, value in enumerate(row):
            ws_cingting.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    ##################################
    for col_idx, col_name in enumerate(cingting_grouped.columns, start=13):  # M 欄開始
        cell=ws_cingting.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    # 將 df_grouped 輸出到 M 到 O 欄
    for r_idx, row in cingting_grouped.iterrows():
        for c_idx, value in enumerate(row):
            ws_cingting.cell(row=r_idx + 2, column=c_idx + 13, value=value)
    #############################
    for col_idx, col_name in enumerate(cingting_day.columns, start=18):  
        cell=ws_cingting.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    for r_idx, row in cingting_day.iterrows():
        for c_idx, value in enumerate(row):
            ws_cingting.cell(row=r_idx + 2, column=c_idx + 18, value=value)

    #######################全日###############################
    ws_changl = wb.create_sheet("王座全日明細")
    # 在 A 到 K 欄顯示 df 的欄位名稱和數據
    for col_idx, col_name in enumerate(changl_o.columns, start=1):
        cell=ws_changl.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    # 將 df 輸出到 A 到 K 欄
    for r_idx, row in changl_o.iterrows():
        for c_idx, value in enumerate(row):
            ws_changl.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    
    #############################
    for col_idx, col_name in enumerate(changl_day.columns, start=21):  
        cell=ws_changl.cell(row=1, column=col_idx, value=col_name)  # 放置欄位名稱
        cell.fill = header_fill  
    for r_idx, row in changl_day.iterrows():
        for c_idx, value in enumerate(row):
            ws_changl.cell(row=r_idx + 2, column=c_idx + 21, value=value)
    ##############調整欄寬#################
    for col in ws.columns:
        max_length = 11
        col_letter = get_column_letter(col[0].column)  # 获取列字母

        # 遍历整列中的所有单元格
        for cell in ws[col_letter]:
            if cell.value is not None:
                value = str(cell.value)
                max_length = max(max_length, len(value))

        # 设置列宽
        ws.column_dimensions[col_letter].width = max_length + 2
    for col in ws_cingting.columns:
        max_length = 11
        col_letter = get_column_letter(col[0].column)  # 获取列字母

        # 遍历整列中的所有单元格
        for cell in ws_cingting[col_letter]:
            if cell.value is not None:
                value = str(cell.value)
                max_length = max(max_length, len(value))

        # 设置列宽
        ws_cingting.column_dimensions[col_letter].width = max_length + 2
    for col in ws_changl.columns:
        max_length = 11
        col_letter = get_column_letter(col[0].column)  # 获取列字母

        # 遍历整列中的所有单元格
        for cell in ws_changl[col_letter]:
            if cell.value is not None:
                value = str(cell.value)
                max_length = max(max_length, len(value))

        # 设置列宽
        ws_changl.column_dimensions[col_letter].width = max_length + 2
    filename = "王座總表.xlsx"
    full_path = os.path.join(filename)

    
    wb.save(full_path)
    message=full_path
    return message
def search_PURTA(TA002):
    import pyodbc
    import pandas as pd
    from openpyxl import Workbook
    import pandas as pd
    from datetime import datetime
    from openpyxl.styles import PatternFill
    import os
    from openpyxl import Workbook
    from dotenv import load_dotenv
    import os
    ENV = './.env' 
    load_dotenv(dotenv_path=ENV)
    DB_host = os.getenv('DB_host')
    DB_password = os.getenv('DB_password')
    DB_uid=os.getenv('DB_uid')
    DATABASE=os.getenv('DATABASE')
    try:
        # 建立與 SQL Server 的連線
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={DB_host};"  # 替換為完整的伺服器名稱
            f"DATABASE={DATABASE};"
            f"UID={DB_uid};"  # 使用 sa 作為使用者名稱
            f"PWD={DB_password};"  # 替換為 sa 的密碼
            "Trusted_Connection=no;" # 明確使用 SQL Server 認證
              
        )

    

        # SQL1 查詢語句
        Sql1 = f"""
            SELECT TA002
            FROM PURTA
            WHERE TA001='310' AND TA002=?
        """
        df = pd.read_sql(Sql1, conn, params=(TA002,))
       
        conn.close()
        
        if not df.empty:
            return True
        else:
            return False
    except:
        return False
def duplicate_FULL(TA002, date):
    """
    old_ta002: 舊的單號 (例如 '20170309004')
    report_date_str: 格式為 '2025-12-31'
    """
    import pyodbc
    import pandas as pd
    from openpyxl import Workbook
    import pandas as pd
    from datetime import datetime
    from openpyxl.styles import PatternFill
    import os
    from openpyxl import Workbook
    from dotenv import load_dotenv

    import os
    ENV = './.env' 
    load_dotenv(dotenv_path=ENV)
    DB_host = os.getenv('DB_host')
    DB_password = os.getenv('DB_password')
    DB_uid=os.getenv('DB_uid')
    DATABASE=os.getenv('DATABASE')
    # DATABASE='KingzaTest'#測試DB
    # 格式化日期格式為 YYYYMMDD
    clean_date = date.replace('-', '')
    conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={DB_host};"  # 替換為完整的伺服器名稱
            f"DATABASE={DATABASE};"
            f"UID={DB_uid};"  # 使用 sa 作為使用者名稱
            f"PWD={DB_password};"  # 替換為 sa 的密碼
            "Trusted_Connection=no;" # 明確使用 SQL Server 認證
              
        )
    cursor = conn.cursor()
    
    try:
        # 1. 取得新單號 (TA002 / TB002)
        cursor.execute("SELECT MAX(TA002) FROM PURTA WHERE TA002 LIKE ?", (f"{clean_date}%",))
        max_id = cursor.fetchone()[0]
        new_no = f"{clean_date}{int(max_id[-3:]) + 1:03d}" if max_id else f"{clean_date}001"

        # --- 開始複製 PURTA (單頭) ---
        insert_ta_sql = f"""
        INSERT INTO PURTA ([COMPANY], [CREATOR], [USR_GROUP], [CREATE_DATE], [MODIFIER], [MODI_DATE], [FLAG], 
            [CREATE_TIME], [MODI_TIME], [TRANS_TYPE], [TRANS_NAME], [TA001], [TA002], [TA003], [TA006], 
            [TA007], [TA008], [TA009], [TA013], [TA014], [TA016], [TA017], 
            [TA004], [TA005], [TA010], [TA011], [TA012], [TA015], [TA018], [TA019], [TA020], [TA021], [TA022], 
            [TA023], [TA024], [TA025], [TA026], [TA027], [TA028], [TA029], [TA030], [TA031], [TA032], [TA033], 
            [TA034], [TA035], [TA036], [TA037], [TA038], [TA039], [TA040], [TA041], [TA042], [TA043], [TA044], 
            [TA045], [TA046], [UDF01], [UDF02], [UDF03], [UDF04], [UDF05], [UDF06], [UDF07], [UDF08], [UDF09], [UDF10])
        SELECT [COMPANY], [CREATOR], [USR_GROUP], '{clean_date}', '', '', '1', '01:00:00', '', [TRANS_TYPE], 
            'newretail_dupli', [TA001], '{new_no}', '{clean_date}', '', 'N', '0', '0', '{clean_date}', '', 'N', '0',
            [TA004], [TA005], [TA010], [TA011], [TA012], [TA015], [TA018], [TA019], [TA020], [TA021], [TA022], 
            [TA023], [TA024], [TA025], [TA026], [TA027], [TA028], [TA029], [TA030], [TA031], [TA032], [TA033], 
            [TA034], [TA035], [TA036], [TA037], [TA038], [TA039], [TA040], [TA041], [TA042], [TA043], [TA044], 
            [TA045], [TA046], [UDF01], [UDF02], [UDF03], [UDF04], [UDF05], [UDF06], [UDF07], [UDF08], [UDF09], [UDF10]
        FROM PURTA WHERE TA002 = ?
        """
        cursor.execute(insert_ta_sql, (TA002,))

        # --- 開始複製 PURTB (單身) ---
        # 注意：TB018 = TB014 * TB017
        insert_tb_sql = f"""
            INSERT INTO PURTB ([COMPANY], [CREATOR], [USR_GROUP], [CREATE_DATE], [MODIFIER], [MODI_DATE], [FLAG], 
                [CREATE_TIME], [MODI_TIME], [TRANS_TYPE], [TRANS_NAME], [TB001], [TB002], [TB011], [TB013], 
                [TB014], [TB015], [TB018], [TB019], [TB020], [TB021], [TB022], [TB025],
                [TB003], [TB004], [TB005], [TB006], [TB007], [TB008], [TB009], [TB010], [TB012], [TB016], [TB017], 
                -- ... 其餘欄位不變 ...
                [TB023], [TB024], [TB026], [TB027], [TB028], [TB029], [TB030], [TB031], [TB032], [TB033], [TB034], 
                [TB035], [TB036], [TB037], [TB038], [TB039], [TB040], [TB041], [TB042], [TB043], [TB044], [TB045], 
                [TB046], [TB047], [TB048], [TB049], [TB050], [TB051], [TB052], [TB053], [TB054], [TB055], [TB056], 
                [TB057], [TB058], [TB059], [TB060], [TB061], [TB062], [TB063], [TB064], [TB065], [TB066], [TB067], 
                [TB068], [TB069], [TB070], [TB071], [TB072], [TB073], [TB074], [TB075], [TB076], [TB077], [TB078], 
                [TB079], [TB080], [TB081], [TB082], [TB083], [TB084], [TB085], [TB086], [TB087], [TB088], [TB089], 
                [TB090], [TB091], [TB092], [TB093], [TB094], [TB095], [TB096], [TB097], [TB098], [TB099], 
                [UDF01], [UDF02], [UDF03], [UDF04], [UDF05], [UDF06], [UDF07], [UDF08], [UDF09], [UDF10])
            SELECT [COMPANY], [CREATOR], [USR_GROUP], '{clean_date}', '', '', '1', '01:00:00', '', [TRANS_TYPE], 
                'newretail_dupli', [TB001], '{new_no}', '{clean_date}', '', 
                -- TB014 = TB009 (加入 TRY_CAST 預防萬一)
                [TB009], 
                -- TB015 = TB007
                [TB007], 
                -- TB018 = TB007 * TB017 (使用 TRY_CAST 處理垃圾字元，失敗則給 0)
                CAST(
                    ISNULL(TRY_CAST(REPLACE([TB014], ',', '') AS DECIMAL(18, 4)), 0) * ISNULL(TRY_CAST(REPLACE([TB017], ',', '') AS DECIMAL(18, 4)), 0) 
                AS NVARCHAR(40)),
                '{clean_date}', 'N', 'N', '', 'N',
                -- 下面這些是原封不動複製的欄位
                [TB003], [TB004], [TB005], [TB006], [TB007], [TB008], [TB009], [TB010], [TB012], [TB016], [TB017], 
                -- ... 其餘 UDF 與 TB 欄位比照辦理 ...
                [TB023], [TB024], [TB026], [TB027], [TB028], [TB029], [TB030], [TB031], [TB032], [TB033], [TB034], 
                [TB035], [TB036], [TB037], [TB038], [TB039], [TB040], [TB041], [TB042], [TB043], [TB044], [TB045], 
                [TB046], [TB047], [TB048], [TB049], [TB050], [TB051], [TB052], [TB053], [TB054], [TB055], [TB056], 
                [TB057], [TB058], [TB059], [TB060], [TB061], [TB062], [TB063], [TB064], [TB065], [TB066], [TB067], 
                [TB068], [TB069], [TB070], [TB071], [TB072], [TB073], [TB074], [TB075], [TB076], [TB077], [TB078], 
                [TB079], [TB080], [TB081], [TB082], [TB083], [TB084], [TB085], [TB086], [TB087], [TB088], [TB089], 
                [TB090], [TB091], [TB092], [TB093], [TB094], [TB095], [TB096], [TB097], [TB098], [TB099], 
                [UDF01], [UDF02], [UDF03], [UDF04], [UDF05], [UDF06], [UDF07], [UDF08], [UDF09], [UDF10]
            FROM PURTB WHERE TB002 = ?
        """
        cursor.execute(insert_tb_sql, (TA002,))

        conn.commit()
        return True, new_no
        
    except Exception as e:
        conn.rollback()
        return False, str(e)
    finally:
        conn.close()
def get_format(productno):
    import os
    from dotenv import load_dotenv
    import pyodbc
    ENV = './.env' 
    load_dotenv(dotenv_path=ENV)
    DB_host = os.getenv('DB_host')
    DB_password = os.getenv('DB_password')
    DB_uid=os.getenv('DB_uid')
    DATABASE=os.getenv('DATABASE')
    # DATABASE='KingzaTest'#測試DB
    # 格式化日期格式為 YYYYMMDD
    try:
        conn = pyodbc.connect(
                "DRIVER={ODBC Driver 17 for SQL Server};"
                f"SERVER={DB_host};"  # 替換為完整的伺服器名稱
                f"DATABASE={DATABASE};"
                f"UID={DB_uid};"  # 使用 sa 作為使用者名稱
                f"PWD={DB_password};"  # 替換為 sa 的密碼
                "Trusted_Connection=no;" # 明確使用 SQL Server 認證
                
            )
        cursor = conn.cursor()
        sele_sql= f""" SELECT [MB003] FROM INVMB WHERE MB001 = ? """
        cursor.execute(sele_sql, (productno,))
        row = cursor.fetchone()
        if row:
            product_name = row[0] # 取得 MB003 的值
            return True, product_name
        else:
            return False, ""
        
    except Exception as e:
        conn.rollback()
        return False, str(e)
    finally:
        conn.close()

# print(duplicate_FULL('20170309001','2025-12-31'))